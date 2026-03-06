# reader_databricks_health.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Dict, Optional, Tuple, List

import re
import pandas as pd
from openpyxl import load_workbook

import config_health as cfg


@dataclass
class DatabricksContext:
    workbook_path: str

    # SSOT header (01_Advertiser_Name)
    hash_name: str
    tenant_id: str
    account_id: str
    downloaded_dt: Optional[datetime]

    # Export downloaded timestamp anchor (date only)
    ref_date: Optional[date]

    # Derived eval window (SSOT from header Date Range)
    window_start: Optional[date]
    window_end: Optional[date]
    window_days: Optional[int]
    window_str: str

    # Backward-compat
    account_name: str

    # sheet_name -> DataFrame
    sheets: Dict[str, pd.DataFrame]

    # inputs / derived
    user_budget: Optional[float] = None

    # constraints/context (read from CS repo)
    primary_kpi: str = "ACOS"
    acos_constraint: Optional[float] = None
    tacos_constraint: Optional[float] = None
    budget_target_from_cs: Optional[float] = None
    season_months: Optional[set[int]] = None
    mrr_fee: Optional[float] = None


def _norm(s: str) -> str:
    return str(s).strip().lower().replace("\n", " ").replace("\r", " ")


def _parse_datetime_any(x) -> Optional[datetime]:
    if x is None:
        return None
    if isinstance(x, datetime):
        return x
    if isinstance(x, date):
        return datetime(x.year, x.month, x.day)
    s = str(x).strip()
    if not s:
        return None

    # common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass

    # fallback: try pandas
    try:
        dt = pd.to_datetime(s, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.to_pydatetime()
    except Exception:
        return None


def _parse_date_any(x) -> Optional[date]:
    dt = _parse_datetime_any(x)
    return dt.date() if dt else None


def _safe_cell(ws, row: int, col: int):
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


def _clean_hash_name(a1_value: str) -> str:
    """
    Databricks often appends a suffix like:
      " - Advertiser_Name"
      " - Advertiser Name"
    Sometimes with underscores/spaces variations.
    We remove that suffix and keep the FULL account string.
    """
    s = str(a1_value or "").strip()

    # Remove any trailing suffix that contains "Advertiser" and "Name" (underscore/space tolerant)
    # Examples removed:
    #   " - Advertiser_Name"
    #   " - Advertiser Name"
    #   "-Advertiser_Name"
    s = re.sub(r"\s*-\s*Advertiser[_\s]*Name\s*$", "", s, flags=re.I)

    # Also handle the case where only " - Advertiser" is appended (rare)
    s = re.sub(r"\s*-\s*Advertiser\s*$", "", s, flags=re.I)

    return s.strip()


def _extract_header_from_01(
    wb_path: str
) -> Tuple[str, str, str, Optional[datetime], str, Optional[date], Optional[date], Optional[int]]:
    """
    Reads SSOT header from 01_Advertiser_Name. This matches the Framework agent logic.
    Expected:
      A1: "<HASH> - ... - Advertiser_Name" (or similar)
      A2: "Tenant ID: ..."
      A3: "Account ID: ..."
      A4: "Date Range: YYYY-MM-DD to YYYY-MM-DD"
      A5: "Downloaded: YYYY-MM-DD HH:MM:SS"
    """
    wb = load_workbook(wb_path, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if name.startswith("01_Advertiser_Name"):
            sheet = wb[name]
            break
    if sheet is None:
        return ("", "", "", None, "", None, None, None)

    # ✅ FIX: Keep FULL account name, only remove the trailing Advertiser suffix
    a1 = _safe_cell(sheet, 1, 1) or ""
    hash_name = _clean_hash_name(a1)

    tenant_id = ""
    account_id = ""
    window_str = ""
    window_start = None
    window_end = None
    window_days = None
    downloaded_dt = None

    # scan first ~20 rows for known patterns
    for r in range(1, 25):
        v = _safe_cell(sheet, r, 1)
        if not v:
            continue
        s = str(v).strip()

        if "tenant id" in s.lower():
            m = re.search(r"tenant id:\s*(.*)$", s, flags=re.I)
            if m:
                tenant_id = m.group(1).strip()

        if "account id" in s.lower():
            m = re.search(r"account id:\s*(.*)$", s, flags=re.I)
            if m:
                account_id = m.group(1).strip()

        if "date range" in s.lower():
            # "Date Range: YYYY-MM-DD to YYYY-MM-DD"
            m = re.search(r"date range:\s*([0-9\-]+)\s*to\s*([0-9\-]+)", s, flags=re.I)
            if m:
                window_start = _parse_date_any(m.group(1))
                window_end = _parse_date_any(m.group(2))
                window_str = f"{m.group(1)} to {m.group(2)}"
                if window_start and window_end:
                    window_days = (window_end - window_start).days

        if "downloaded" in s.lower():
            m = re.search(r"downloaded:\s*(.*)$", s, flags=re.I)
            if m:
                downloaded_dt = _parse_datetime_any(m.group(1).strip())

    return (hash_name, tenant_id, account_id, downloaded_dt, window_str, window_start, window_end, window_days)


def _load_allowed_sheets_to_dfs(wb_path: str) -> Dict[str, pd.DataFrame]:
    """
    Mirrors Framework agent: load only sheet names that match allowlist prefixes.
    Use header row 6 (0-index header=5) for Databricks exports.
    """
    xl = pd.ExcelFile(wb_path)
    sheet_names = xl.sheet_names

    # build allowlist prefixes from cfg.TAB_CANDIDATES
    allowed_prefixes: List[str] = []
    for prefixes in cfg.TAB_CANDIDATES.values():
        allowed_prefixes.extend(prefixes)

    sheets: Dict[str, pd.DataFrame] = {}
    for s in sheet_names:
        if any(s.startswith(pfx) for pfx in allowed_prefixes):
            try:
                df = pd.read_excel(wb_path, sheet_name=s, header=5)
                sheets[s] = df
            except Exception:
                # keep going; missing/failed sheets will be flagged at control evaluation
                continue
    return sheets


def load_databricks_context(workbook_path: str) -> DatabricksContext:
    hash_name, tenant_id, account_id, downloaded_dt, window_str, window_start, window_end, window_days = _extract_header_from_01(workbook_path)
    ref_date = downloaded_dt.date() if downloaded_dt else None

    sheets = _load_allowed_sheets_to_dfs(workbook_path)

    # account_name kept for compatibility (use hash_name)
    account_name = hash_name

    return DatabricksContext(
        workbook_path=workbook_path,
        hash_name=hash_name,
        tenant_id=tenant_id,
        account_id=account_id,
        downloaded_dt=downloaded_dt,
        ref_date=ref_date,
        window_start=window_start,
        window_end=window_end,
        window_days=window_days,
        window_str=window_str,
        account_name=account_name,
        sheets=sheets,
        season_months=set(),
    )


def get_dataset(ctx: DatabricksContext, dataset_key: str) -> Optional[pd.DataFrame]:
    prefixes = cfg.TAB_CANDIDATES.get(dataset_key, [])
    for sheet_name, df in ctx.sheets.items():
        if any(sheet_name.startswith(pfx) for pfx in prefixes):
            return df
    return None
