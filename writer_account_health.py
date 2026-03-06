# writer_account_health.py
from __future__ import annotations

from typing import Dict
from openpyxl import load_workbook

import config_health as cfg
from reader_databricks_health import DatabricksContext


def write_account_health_output(
    template_path: str,
    output_path: str,
    ctx: DatabricksContext,
    results: Dict[str, cfg.ControlResult],
) -> None:
    # keep_vba=True is required to preserve macros
    wb = load_workbook(template_path, keep_vba=True)

    ws_main = wb["Account Health_Analysis"]
    ws_ref = wb["Account Health_Reference"]

    # ---- Header/meta (mirrors Framework writer approach) ----
    ws_main["A1"].value = f"{ctx.hash_name} — Account Health Analysis"
    ws_main["B3"].value = f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}"

    if ctx.window_start and ctx.window_end and ctx.window_days is not None:
        ws_main["B4"].value = f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)"
    else:
        ws_main["B4"].value = ctx.window_str or ""

    ws_main["B5"].value = ctx.downloaded_dt.strftime("%Y-%m-%d %H:%M:%S") if ctx.downloaded_dt else ""

    # ---- Constraints summary block (your requested cells) ----
    ws_main["B9"].value = f"{ctx.acos_constraint*100:.1f}%" if ctx.acos_constraint is not None else "NOT FOUND"
    ws_main["B10"].value = f"{ctx.tacos_constraint*100:.1f}%" if ctx.tacos_constraint is not None else "NOT FOUND"
    ws_main["B11"].value = float(ctx.budget_target_from_cs) if ctx.budget_target_from_cs is not None else "NOT DEFINED"
    ws_main["E10"].value = ctx.primary_kpi or "ACOS"

    # ---- Write control results to Account Health_Reference ----
    # Column mapping (per your template)
    control_id_col = 2   # B
    status_col = 4       # D
    what_col = 8         # H
    why_col = 9          # I
    src_col = 10         # J

    # ------------------------------------------------------------------
    # SPEED UPGRADE:
    # 1) Build a row index map once (ControlID -> row)
    # 2) Write by iterating over `results` (not scanning the sheet repeatedly)
    # This is notably faster when scaling or when templates get larger.
    # ------------------------------------------------------------------
    cid_to_row: Dict[str, int] = {}

    # Only scan the Control ID column once
    for r in range(2, ws_ref.max_row + 1):
        cid = ws_ref.cell(row=r, column=control_id_col).value
        if cid:
            cid_to_row[str(cid).strip().upper()] = r

    # Now write results directly
    for cid, res in results.items():
        r = cid_to_row.get(str(cid).strip().upper())
        if not r:
            continue

        ws_ref.cell(row=r, column=status_col).value = res.status
        ws_ref.cell(row=r, column=what_col).value = res.what_we_saw
        ws_ref.cell(row=r, column=why_col).value = res.why_it_matters
        ws_ref.cell(row=r, column=src_col).value = res.data_source

    # ------------------------------------------------------------------
    # RELIABILITY UPGRADE:
    # Save + explicit close to release file handle and prevent stale links.
    # ------------------------------------------------------------------
    wb.save(output_path)
    try:
        wb.close()
    except Exception:
        pass
